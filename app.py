from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, date
import os
import io

app = Flask(__name__)

# Load master data
MASTER_FILE = "RSSB Workflow Final.xlsx"
TEMPLATE_FILE = "NR_May 2026 Construction Beas.xlsx"
SEWA_HISTORY_FILE = "sewa_history_log.xlsx"

# Cache master data in memory for faster searches
_master_data_cache = None

def load_master_data():
    """Load master data from Excel file (cached for performance)"""
    global _master_data_cache
    if _master_data_cache is None:
        print("Loading master data from Excel file...")
        _master_data_cache = pd.read_excel(MASTER_FILE, sheet_name="Planner Persons")
        print(f"Master data loaded: {len(_master_data_cache)} records")
    return _master_data_cache

def reload_master_data():
    """Force reload master data from Excel file"""
    global _master_data_cache
    _master_data_cache = None
    return load_master_data()

def search_names(query):
    """Search for names in master data"""
    df = load_master_data()
    # Search in Name column (case-insensitive)
    matches = df[df['Name'].str.contains(query, case=False, na=False)]
    
    # Return list of matching names with their details
    results = []
    for _, row in matches.iterrows():
        # Safely convert age to int, handle errors
        try:
            age = int(row['Age']) if pd.notna(row['Age']) and str(row['Age']).replace('.','').isdigit() else ''
        except (ValueError, TypeError):
            age = ''
        
        results.append({
            'name': row['Name'],
            'father_name': row['Father/ Husband/ Mother name'] if pd.notna(row['Father/ Husband/ Mother name']) else '',
            'age': age,
            'gender': row['Gender'] if pd.notna(row['Gender']) else '',
            'contact': str(row['Contact No']) if pd.notna(row['Contact No']) else '',
            'badge': row['Badge No (Centre)'] if pd.notna(row['Badge No (Centre)']) else 'NA',
            'aadhar': str(row['Aadhar']) if pd.notna(row['Aadhar']) else '',
            'address': row['Address (current)'] if pd.notna(row['Address (current)']) else ''
        })
    return results

def normalize_selected_names(selected_names, jathedar_name=''):
    """Ensure selected names are unique and jathedar stays first when selected"""
    ordered_names = []
    seen = set()

    if jathedar_name and jathedar_name not in seen:
        ordered_names.append(jathedar_name)
        seen.add(jathedar_name)

    for name in selected_names:
        if name and name not in seen:
            ordered_names.append(name)
            seen.add(name)

    return ordered_names


def save_to_sewa_history(selected_names, metadata):
    """Save grouped nominal roll history to Excel log file"""
    try:
        try:
            history_df = pd.read_excel(SEWA_HISTORY_FILE, sheet_name='Sewa History')
        except FileNotFoundError:
            history_df = pd.DataFrame(columns=[
                'NR_ID', 'Timestamp', 'Reference_No', 'Sewadar_Name',
                'Selected_Names', 'Place_of_Sewa', 'From_Date', 'To_Date',
                'Duration_Days', 'Jathedar', 'Driver',
                'Vehicle_Type', 'Vehicle_No'
            ])

        try:
            from_date = datetime.strptime(metadata.get('from_date', ''), '%Y-%m-%d')
            to_date = datetime.strptime(metadata.get('to_date', ''), '%Y-%m-%d')
            duration = (to_date - from_date).days + 1
        except:
            duration = 0

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        nr_id = f"NR-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        normalized_names = normalize_selected_names(selected_names, metadata.get('jathedar', ''))
        joined_names = " | ".join(normalized_names)

        new_records = []
        for name in normalized_names:
            new_records.append({
                'NR_ID': nr_id,
                'Timestamp': timestamp,
                'Reference_No': metadata.get('reference_no', ''),
                'Sewadar_Name': name,
                'Selected_Names': joined_names,
                'Place_of_Sewa': metadata.get('place_of_sewa', ''),
                'From_Date': metadata.get('from_date', ''),
                'To_Date': metadata.get('to_date', ''),
                'Duration_Days': duration,
                'Jathedar': metadata.get('jathedar', ''),
                'Driver': metadata.get('driver', ''),
                'Vehicle_Type': metadata.get('vehicle_type', ''),
                'Vehicle_No': metadata.get('vehicle_no', '')
            })

        if new_records:
            history_df = pd.concat([history_df, pd.DataFrame(new_records)], ignore_index=True)
            history_df.to_excel(SEWA_HISTORY_FILE, index=False, sheet_name='Sewa History')

        return True
    except Exception as e:
        print(f"Error saving to sewa history: {e}")
        return False

def mask_aadhar(aadhar):
    """Mask Aadhar number to show only last 4 digits"""
    aadhar_str = str(aadhar)
    if len(aadhar_str) >= 4:
        return '*' * (len(aadhar_str) - 4) + aadhar_str[-4:]
    return aadhar_str


def load_history_dataframe():
    """Load sewa history as dataframe"""
    try:
        history_df = pd.read_excel(SEWA_HISTORY_FILE, sheet_name='Sewa History')
    except FileNotFoundError:
        history_df = pd.DataFrame(columns=[
            'NR_ID', 'Timestamp', 'Reference_No', 'Sewadar_Name',
            'Selected_Names', 'Place_of_Sewa', 'From_Date', 'To_Date',
            'Duration_Days', 'Jathedar', 'Driver', 'Vehicle_Type', 'Vehicle_No'
        ])
    return history_df


def sanitize_records(records):
    """Convert pandas/numpy values to JSON-safe Python values"""
    cleaned_records = []

    def sanitize_value(value):
        if isinstance(value, list):
            return [sanitize_value(item) for item in value]
        if isinstance(value, tuple):
            return [sanitize_value(item) for item in value]
        if isinstance(value, dict):
            return {key: sanitize_value(val) for key, val in value.items()}
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        return value

    for record in records:
        cleaned_record = {}
        for key, value in record.items():
            cleaned_record[key] = sanitize_value(value)
        cleaned_records.append(cleaned_record)
    return cleaned_records


def build_nr_groups(history_df):
    """Build grouped NR records from per-sewadar history rows"""
    if history_df.empty:
        return []

    def clean_scalar(value, default=''):
        if isinstance(value, (list, tuple)):
            return default
        if value is None:
            return default
        try:
            if pd.isna(value):
                return default
        except Exception:
            pass
        value_str = str(value).strip()
        return default if value_str.lower() == 'nan' else value_str

    history_df = history_df.copy()

    required_columns = [
        'Timestamp', 'Reference_No', 'Sewadar_Name', 'Selected_Names', 'Place_of_Sewa',
        'From_Date', 'To_Date', 'Jathedar', 'Driver', 'Vehicle_Type', 'Vehicle_No',
        'Duration_Days', 'NR_ID'
    ]
    for column in required_columns:
        if column not in history_df.columns:
            history_df[column] = ''

    history_df['Timestamp'] = history_df['Timestamp'].apply(clean_scalar)
    history_df['Reference_No'] = history_df['Reference_No'].apply(clean_scalar)
    history_df['Sewadar_Name'] = history_df['Sewadar_Name'].apply(clean_scalar)
    history_df['Selected_Names'] = history_df['Selected_Names'].apply(clean_scalar)
    history_df['Place_of_Sewa'] = history_df['Place_of_Sewa'].apply(clean_scalar)
    history_df['From_Date'] = history_df['From_Date'].apply(clean_scalar)
    history_df['To_Date'] = history_df['To_Date'].apply(clean_scalar)
    history_df['Jathedar'] = history_df['Jathedar'].apply(clean_scalar)
    history_df['Driver'] = history_df['Driver'].apply(clean_scalar)
    history_df['Vehicle_Type'] = history_df['Vehicle_Type'].apply(clean_scalar)
    history_df['Vehicle_No'] = history_df['Vehicle_No'].apply(clean_scalar)
    history_df['NR_ID'] = history_df['NR_ID'].apply(clean_scalar)

    missing_nr_ids = history_df['NR_ID'] == ''
    history_df.loc[missing_nr_ids, 'NR_ID'] = (
        history_df.loc[missing_nr_ids, 'Timestamp'].astype(str) + '_' +
        history_df.loc[missing_nr_ids, 'Reference_No'].astype(str)
    )

    grouped_records = []
    today = pd.to_datetime(date.today())

    for nr_id, group in history_df.groupby('NR_ID', dropna=False, sort=False):
        if group.empty:
            continue

        first_row = group.iloc[0]
        from_date_raw = clean_scalar(first_row.get('From_Date', ''))
        from_date_value = pd.to_datetime(from_date_raw, errors='coerce')
        status = 'upcoming' if pd.notna(from_date_value) and from_date_value >= today else 'past'

        selected_names_str = clean_scalar(first_row.get('Selected_Names', ''))
        if selected_names_str:
            selected_name_list = [name.strip() for name in selected_names_str.split(' | ') if name.strip()]
        else:
            selected_name_list = []
            for name in group['Sewadar_Name'].tolist():
                clean_name = clean_scalar(name)
                if clean_name and clean_name not in selected_name_list:
                    selected_name_list.append(clean_name)
            selected_names_str = " | ".join(selected_name_list)

        duration_value = first_row.get('Duration_Days', 0)
        try:
            if pd.isna(duration_value):
                duration_value = 0
        except Exception:
            duration_value = 0

        try:
            duration_value = int(float(duration_value))
        except (TypeError, ValueError):
            duration_value = 0

        grouped_records.append({
            'NR_ID': clean_scalar(nr_id),
            'Timestamp': clean_scalar(first_row.get('Timestamp', '')),
            'Reference_No': clean_scalar(first_row.get('Reference_No', '')),
            'Jathedar': clean_scalar(first_row.get('Jathedar', '')),
            'Driver': clean_scalar(first_row.get('Driver', '')),
            'Vehicle_Type': clean_scalar(first_row.get('Vehicle_Type', '')),
            'Vehicle_No': clean_scalar(first_row.get('Vehicle_No', '')),
            'Place_of_Sewa': clean_scalar(first_row.get('Place_of_Sewa', '')),
            'From_Date': from_date_raw,
            'To_Date': clean_scalar(first_row.get('To_Date', '')),
            'Duration_Days': duration_value,
            'Selected_Names': selected_names_str,
            'Selected_Name_List': selected_name_list,
            'Sewadar_Count': len(selected_name_list),
            'Status': status,
            'Can_Edit': status == 'upcoming'
        })

    grouped_records.sort(key=lambda record: record.get('Timestamp', ''), reverse=True)
    return grouped_records

def generate_nominal_roll(selected_names, metadata):
    """Generate nominal roll Excel file"""
    selected_names = normalize_selected_names(selected_names, metadata.get('jathedar', ''))

    def format_display_date(value):
        if not value:
            return ''
        try:
            return datetime.strptime(str(value), '%Y-%m-%d').strftime('%d-%m-%Y')
        except ValueError:
            return str(value)

    # Load template with data_only=True to convert formulas to values
    wb = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
    ws = wb.active
    
    # Remove all sheets except the active one
    sheets_to_remove = [sheet for sheet in wb.sheetnames if sheet != ws.title]
    for sheet_name in sheets_to_remove:
        del wb[sheet_name]
    
    # Convert all formulas to values in the remaining sheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # Clear formula cells that reference deleted sheets
                cell.value = None
    
    # Update reference number (row 3)
    ws['A3'] = metadata.get('reference_no', '')
    
    # Update metadata in the template with hardcoded values
    ws['C5'] = 'Indirapuram'  # Hardcoded Satsang Place
    ws['G5'] = 'Ghaziabad'    # Hardcoded Area
    ws['I5'] = 'III'          # Hardcoded Zone
    ws['C6'] = metadata.get('jathedar', '')
    ws['G6'] = metadata.get('driver', '')
    ws['C7'] = metadata.get('vehicle_type', '')
    ws['G7'] = metadata.get('vehicle_no', '')
    ws['C8'] = metadata.get('place_of_sewa', '')
    ws['G8'] = format_display_date(metadata.get('from_date', ''))
    ws['I8'] = format_display_date(metadata.get('to_date', ''))
    
    # Load master data
    df = load_master_data()
    
    # Starting row for data (after header row 11)
    start_row = 12
    
    # First, delete all existing data rows from the template (row 12 onwards)
    # Get the max row in the template
    max_row = ws.max_row
    if max_row >= start_row:
        ws.delete_rows(start_row, max_row - start_row + 1)
    
    # Add selected sewadars
    for idx, name in enumerate(selected_names):
        row_num = start_row + idx
        
        # Find person in master data
        person = df[df['Name'] == name]
        if person.empty:
            continue
            
        person = person.iloc[0]
        
        # Fill data in the row
        ws[f'A{row_num}'] = idx + 1
        ws[f'B{row_num}'] = person['Name']
        ws[f'C{row_num}'] = person['Father/ Husband/ Mother name'] if pd.notna(person['Father/ Husband/ Mother name']) else ''
        ws[f'D{row_num}'] = person['Gender'] if pd.notna(person['Gender']) else ''
        ws[f'E{row_num}'] = int(person['Age']) if pd.notna(person['Age']) else ''
        ws[f'F{row_num}'] = mask_aadhar(person['Aadhar']) if pd.notna(person['Aadhar']) else ''
        ws[f'G{row_num}'] = person['Address (current)'] if pd.notna(person['Address (current)']) else ''
        ws[f'H{row_num}'] = str(person['Contact No']) if pd.notna(person['Contact No']) else ''
        ws[f'I{row_num}'] = person['Badge No (Centre)'] if pd.notna(person['Badge No (Centre)']) else 'NA'
        
        # Apply formatting to match template - bold text with borders
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row_num}']
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Calculate the next row after all sewadar data
    last_data_row = start_row + len(selected_names)
    
    # Add 4 empty rows for manual entries
    for i in range(4):
        row_num = last_data_row + i
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row_num}']
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Add blank space rows (7 rows with borders around the section)
    blank_space_start = last_data_row + 4
    blank_space_rows = 7

    # Merge the full pre-footer blank area row-wise and keep only the outer border
    for i in range(blank_space_rows):
        row_num = blank_space_start + i
        ws.merge_cells(f'A{row_num}:I{row_num}')

        merged_cell = ws[f'A{row_num}']
        merged_cell.font = Font(name='Calibri', size=12, bold=True)
        merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row_num}']
            cell.border = Border(
                left=Side(style='thin') if col == 'A' else None,
                right=Side(style='thin') if col == 'I' else None,
                top=Side(style='thin') if i == 0 else None,
                bottom=Side(style='thin') if i == blank_space_rows - 1 else None
            )
    
    # Add footer section matching the reference layout
    footer_start_row = blank_space_start + blank_space_rows  # After blank space rows

    # Merge footer cells to create left and right blocks
    ws.merge_cells(f'A{footer_start_row}:D{footer_start_row}')
    ws.merge_cells(f'F{footer_start_row}:I{footer_start_row}')
    ws.merge_cells(f'A{footer_start_row + 1}:D{footer_start_row + 1}')
    ws.merge_cells(f'F{footer_start_row + 1}:I{footer_start_row + 1}')
    ws.merge_cells(f'A{footer_start_row + 2}:D{footer_start_row + 2}')
    ws.merge_cells(f'F{footer_start_row + 2}:I{footer_start_row + 2}')
    ws.merge_cells(f'A{footer_start_row + 3}:D{footer_start_row + 3}')
    ws.merge_cells(f'F{footer_start_row + 3}:I{footer_start_row + 3}')

    # Footer text
    ws[f'A{footer_start_row}'] = '(Signature of Jathedar)'
    ws[f'F{footer_start_row}'] = '(Signature of Functionary)'
    ws[f'A{footer_start_row + 1}'] = '(Affixx Rubber Stamp)'
    ws[f'F{footer_start_row + 1}'] = '(Affixx Rubber Stamp)'
    ws[f'A{footer_start_row + 2}'] = 'Date :'
    ws[f'F{footer_start_row + 2}'] = 'Date :'
    ws[f'A{footer_start_row + 3}'] = 'Contact No.   :'
    ws[f'F{footer_start_row + 3}'] = 'Contact No.   :'

    # Format footer
    for row_offset in range(4):
        row_num = footer_start_row + row_offset
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row_num}']
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Set row heights for print-ready larger font and proper wrapping
    for row in range(1, 12):
        ws.row_dimensions[row].height = 24

    for row in range(start_row, last_data_row):
        ws.row_dimensions[row].height = 36

    for row in range(last_data_row, blank_space_start):
        ws.row_dimensions[row].height = 36

    for row in range(blank_space_start, footer_start_row):
        ws.row_dimensions[row].height = 28

    ws.row_dimensions[footer_start_row].height = 34
    ws.row_dimensions[footer_start_row + 1].height = 34
    ws.row_dimensions[footer_start_row + 2].height = 24
    ws.row_dimensions[footer_start_row + 3].height = 24

    # Apply Calibri 34 bold formatting to all header rows and labels
    for row in range(1, 12):  # Rows 1-11 (headers and labels)
        for col in range(1, 10):  # Columns A-I
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Keep body/footer rows left-aligned while preserving centered header rows
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        for row in range(12, footer_start_row + 4):
            ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Explicitly center the title and reference rows requested by the user
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Expand key columns to reduce text clipping in print-ready output
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14
    
    # Save to BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])
def search():
    query = request.json.get('query', '')
    if len(query) < 2:
        return jsonify([])
    
    results = search_names(query)
    return jsonify(results)

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    metadata = data.get('metadata', {})
    selected_names = normalize_selected_names(data.get('names', []), metadata.get('jathedar', ''))
    save_history = data.get('save_history', True)

    if not selected_names:
        return jsonify({'error': 'No names selected'}), 400

    if save_history:
        save_to_sewa_history(selected_names, metadata)

    output = generate_nominal_roll(selected_names, metadata)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Nominal_Roll_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports')
def reports():
    """Render reports page"""
    return render_template('reports.html')

@app.route('/api/sewa-history', methods=['GET'])
def get_sewa_history():
    """Get sewa history data with optional filters"""
    try:
        history_df = load_history_dataframe()

        # Normalize text/date columns for reliable filtering
        history_df['Sewadar_Name'] = history_df['Sewadar_Name'].fillna('').astype(str)
        history_df['Place_of_Sewa'] = history_df['Place_of_Sewa'].fillna('').astype(str)
        history_df['From_Date_Filter'] = pd.to_datetime(history_df['From_Date'], errors='coerce')
        history_df['To_Date_Filter'] = pd.to_datetime(history_df['To_Date'], errors='coerce')

        # Get filter parameters
        sewadar_name = request.args.get('sewadar_name', '').strip()
        place = request.args.get('place', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()

        # Apply filters
        if sewadar_name:
            history_df = history_df[history_df['Sewadar_Name'].str.contains(sewadar_name, case=False, na=False)]

        if place:
            history_df = history_df[history_df['Place_of_Sewa'].str.contains(place, case=False, na=False)]

        if from_date:
            from_date_value = pd.to_datetime(from_date, errors='coerce')
            if pd.notna(from_date_value):
                history_df = history_df[history_df['From_Date_Filter'].notna() & (history_df['From_Date_Filter'] >= from_date_value)]

        if to_date:
            to_date_value = pd.to_datetime(to_date, errors='coerce')
            if pd.notna(to_date_value):
                history_df = history_df[history_df['To_Date_Filter'].notna() & (history_df['To_Date_Filter'] <= to_date_value)]

        # Remove helper columns and sanitize values for JSON/frontend rendering
        history_df = history_df.drop(columns=['From_Date_Filter', 'To_Date_Filter'], errors='ignore')
        datetime_columns = history_df.select_dtypes(include=['datetime64[ns]', 'datetimetz']).columns
        for column in datetime_columns:
            history_df[column] = history_df[column].astype(str)

        records = sanitize_records(history_df.to_dict('records'))

        return jsonify({
            'success': True,
            'data': records,
            'count': len(records)
        })
    except FileNotFoundError:
        return jsonify({
            'success': True,
            'data': [],
            'count': 0,
            'message': 'No sewa history found yet'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/sewadar-summary/<sewadar_name>', methods=['GET'])
def get_sewadar_summary(sewadar_name):
    """Get summary statistics for a specific sewadar"""
    try:
        history_df = pd.read_excel(SEWA_HISTORY_FILE, sheet_name='Sewa History')
        
        # Filter for specific sewadar
        sewadar_data = history_df[history_df['Sewadar_Name'] == sewadar_name]
        
        if sewadar_data.empty:
            return jsonify({
                'success': True,
                'data': {
                    'total_trips': 0,
                    'total_days': 0,
                    'places_visited': [],
                    'trips': []
                }
            })
        
        # Calculate statistics
        total_trips = len(sewadar_data)
        total_days = sewadar_data['Duration_Days'].sum()
        places_visited = sewadar_data['Place_of_Sewa'].unique().tolist()
        
        # Get trip details
        trips = sewadar_data[['Timestamp', 'Place_of_Sewa', 'From_Date', 'To_Date',
                              'Duration_Days', 'Jathedar', 'Reference_No']].to_dict('records')
        
        return jsonify({
            'success': True,
            'data': {
                'sewadar_name': sewadar_name,
                'total_trips': int(total_trips),
                'total_days': int(total_days),
                'places_visited': places_visited,
                'unique_places_count': len(places_visited),
                'trips': trips
            }
        })
    except FileNotFoundError:
        return jsonify({
            'success': True,
            'data': {
                'total_trips': 0,
                'total_days': 0,
                'places_visited': [],
                'trips': []
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """Get overall statistics"""
    try:
        history_df = load_history_dataframe()

        if history_df.empty:
            raise FileNotFoundError

        history_df = history_df.copy()
        history_df['Sewadar_Name'] = history_df['Sewadar_Name'].fillna('').astype(str)
        history_df['Place_of_Sewa'] = history_df['Place_of_Sewa'].fillna('').astype(str)
        history_df['Duration_Days'] = pd.to_numeric(history_df['Duration_Days'], errors='coerce').fillna(0)
        history_df['Timestamp'] = history_df['Timestamp'].fillna('').astype(str)

        valid_sewadar_names = history_df['Sewadar_Name'].str.strip()
        valid_places = history_df['Place_of_Sewa'].str.strip()

        total_sewadars = valid_sewadar_names[valid_sewadar_names != ''].nunique()
        total_trips = len(history_df[valid_sewadar_names != ''])
        total_days = int(history_df['Duration_Days'].sum())
        unique_places = valid_places[valid_places != ''].nunique()

        top_sewadars = history_df.loc[valid_sewadar_names != '', 'Sewadar_Name'].value_counts().head(10).to_dict()
        top_places = history_df.loc[valid_places != '', 'Place_of_Sewa'].value_counts().head(10).to_dict()

        recent_trips = sanitize_records(
            history_df.sort_values('Timestamp', ascending=False).head(10)[
                ['Timestamp', 'Sewadar_Name', 'Place_of_Sewa', 'From_Date', 'To_Date', 'Duration_Days']
            ].to_dict('records')
        )
        
        return jsonify({
            'success': True,
            'data': {
                'total_sewadars': int(total_sewadars),
                'total_trips': int(total_trips),
                'total_days': int(total_days),
                'unique_places': int(unique_places),
                'top_sewadars': top_sewadars,
                'top_places': top_places,
                'recent_trips': recent_trips
            }
        })
    except FileNotFoundError:
        return jsonify({
            'success': True,
            'data': {
                'total_sewadars': 0,
                'total_trips': 0,
                'total_days': 0,
                'unique_places': 0,
                'top_sewadars': {},
                'top_places': {},
                'recent_trips': []
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/generated-nrs', methods=['GET'])
def get_generated_nrs():
    """Get grouped nominal roll generations with upcoming/past status"""
    try:
        history_df = load_history_dataframe()
        grouped_records = build_nr_groups(history_df)
        return jsonify({
            'success': True,
            'data': sanitize_records(grouped_records),
            'count': len(grouped_records)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/generated-nrs/<nr_id>', methods=['PUT'])
def update_generated_nr(nr_id):
    """Update an upcoming NR group"""
    try:
        history_df = load_history_dataframe()
        if history_df.empty or 'NR_ID' not in history_df.columns:
            return jsonify({'success': False, 'error': 'NR not found'}), 404

        target_rows = history_df['NR_ID'].astype(str) == str(nr_id)
        if not target_rows.any():
            return jsonify({'success': False, 'error': 'NR not found'}), 404

        target_group = history_df[target_rows]
        first_from_date = pd.to_datetime(target_group.iloc[0].get('From_Date', ''), errors='coerce')
        if pd.isna(first_from_date) or first_from_date < pd.to_datetime(date.today()):
            return jsonify({'success': False, 'error': 'Only upcoming NRs can be edited'}), 400

        data = request.json or {}
        metadata = data.get('metadata', {})
        selected_names = normalize_selected_names(data.get('names', []), metadata.get('jathedar', ''))
        joined_names = " | ".join(selected_names)

        history_df = history_df[~target_rows]

        try:
            from_date = datetime.strptime(metadata.get('from_date', ''), '%Y-%m-%d')
            to_date = datetime.strptime(metadata.get('to_date', ''), '%Y-%m-%d')
            duration = (to_date - from_date).days + 1
        except:
            duration = 0

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_records = []
        for name in selected_names:
            new_records.append({
                'NR_ID': nr_id,
                'Timestamp': timestamp,
                'Reference_No': metadata.get('reference_no', ''),
                'Sewadar_Name': name,
                'Selected_Names': joined_names,
                'Place_of_Sewa': metadata.get('place_of_sewa', ''),
                'From_Date': metadata.get('from_date', ''),
                'To_Date': metadata.get('to_date', ''),
                'Duration_Days': duration,
                'Jathedar': metadata.get('jathedar', ''),
                'Driver': metadata.get('driver', ''),
                'Vehicle_Type': metadata.get('vehicle_type', ''),
                'Vehicle_No': metadata.get('vehicle_no', '')
            })

        history_df = pd.concat([history_df, pd.DataFrame(new_records)], ignore_index=True)
        history_df.to_excel(SEWA_HISTORY_FILE, index=False, sheet_name='Sewa History')

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/generated-nrs/<nr_id>', methods=['DELETE'])
def delete_generated_nr(nr_id):
    """Delete an upcoming NR group"""
    try:
        history_df = load_history_dataframe()
        if history_df.empty or 'NR_ID' not in history_df.columns:
            return jsonify({'success': False, 'error': 'NR not found'}), 404

        target_rows = history_df['NR_ID'].astype(str) == str(nr_id)
        if not target_rows.any():
            return jsonify({'success': False, 'error': 'NR not found'}), 404

        target_group = history_df[target_rows]
        first_from_date = pd.to_datetime(target_group.iloc[0].get('From_Date', ''), errors='coerce')
        if pd.isna(first_from_date) or first_from_date < pd.to_datetime(date.today()):
            return jsonify({'success': False, 'error': 'Only upcoming NRs can be deleted'}), 400

        history_df = history_df[~target_rows]
        history_df.to_excel(SEWA_HISTORY_FILE, index=False, sheet_name='Sewa History')

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/export-report', methods=['POST'])
def export_report():
    """Export filtered report to Excel"""
    try:
        data = request.json
        sewadar_name = data.get('sewadar_name', '').strip()
        place = data.get('place', '').strip()
        from_date = data.get('from_date', '').strip()
        to_date = data.get('to_date', '').strip()

        history_df = load_history_dataframe()
        history_df['Sewadar_Name'] = history_df['Sewadar_Name'].fillna('').astype(str)
        history_df['Place_of_Sewa'] = history_df['Place_of_Sewa'].fillna('').astype(str)
        history_df['From_Date_Filter'] = pd.to_datetime(history_df['From_Date'], errors='coerce')
        history_df['To_Date_Filter'] = pd.to_datetime(history_df['To_Date'], errors='coerce')

        if sewadar_name:
            history_df = history_df[history_df['Sewadar_Name'].str.contains(sewadar_name, case=False, na=False)]

        if place:
            history_df = history_df[history_df['Place_of_Sewa'].str.contains(place, case=False, na=False)]

        if from_date:
            from_date_value = pd.to_datetime(from_date, errors='coerce')
            if pd.notna(from_date_value):
                history_df = history_df[history_df['From_Date_Filter'].notna() & (history_df['From_Date_Filter'] >= from_date_value)]

        if to_date:
            to_date_value = pd.to_datetime(to_date, errors='coerce')
            if pd.notna(to_date_value):
                history_df = history_df[history_df['To_Date_Filter'].notna() & (history_df['To_Date_Filter'] <= to_date_value)]

        history_df = history_df.drop(columns=['From_Date_Filter', 'To_Date_Filter'], errors='ignore')

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            history_df.to_excel(writer, sheet_name='Filtered Report', index=False)

        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Sewa_Report_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/reload-master-data', methods=['POST'])
def reload_master_data_endpoint():
    """Reload master data from Excel file"""
    try:
        reload_master_data()
        return jsonify({
            'success': True,
            'message': 'Master data reloaded successfully',
            'record_count': len(_master_data_cache)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    # Allow access from other devices on the network (for mobile access)
    # Use PORT environment variable for production deployment (Render, Heroku, etc.)
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, host='0.0.0.0', port=port)

# Made with Bob
